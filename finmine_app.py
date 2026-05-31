"""
finmine_app.py — FinMine Analytics Engine
Main Streamlit UI.

Run:  streamlit run finmine_app.py
"""

import io
import json
import logging
import zipfile
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
import plotly.subplots as sp
import streamlit as st

import database as db
import scraper
import ml_engine
import ai_engine

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)s  %(name)s — %(message)s")
log = logging.getLogger(__name__)

db.init_db()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="FinMine Analytics Engine",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  [data-testid="stMetricValue"]   { font-size: 1.25rem; }
  .stTabs [data-baseweb="tab"]    { font-size: 0.95rem; }
  code { font-size: 0.8rem; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

# openpyxl style constants
_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
_SUB_FILL   = PatternFill("solid", fgColor="2F5496")
_SUB_FONT   = Font(color="FFFFFF", bold=True)
_SB_FILL    = PatternFill("solid", fgColor="C6EFCE")   # light green — STRONG BUY
_BUY_FILL   = PatternFill("solid", fgColor="DDEBF7")   # light blue  — BUY
_HOLD_FILL  = PatternFill("solid", fgColor="FFEB9C")   # light amber — HOLD


def _label_fill(label: str) -> PatternFill:
    if "STRONG" in label:
        return _SB_FILL
    if label == "BUY":
        return _BUY_FILL
    return _HOLD_FILL


def _safe_pct(val) -> str:
    try:
        v = float(val)
        return f"{v * 100:.1f}%" if not np.isnan(v) else "N/A"
    except (TypeError, ValueError):
        return "N/A"


def _safe_ratio(val) -> str:
    try:
        v = float(val)
        return f"{v:.2f}x" if not np.isnan(v) else "N/A"
    except (TypeError, ValueError):
        return "N/A"


def _price_stats(ticker: str) -> dict:
    """Returns price performance stats dict for one ticker."""
    prices = db.get_prices(ticker)
    out = dict(latest="", w52h="", w52l="", ret1="", ret3="", ret10="")
    if prices.empty:
        return out
    prices = prices.sort_values("date")
    closes = prices["close"].astype(float)
    latest = float(closes.iloc[-1])
    out["latest"] = f"${latest:,.2f}"

    last_date = pd.to_datetime(prices["date"].iloc[-1])

    for years, key in [(1, "ret1"), (3, "ret3"), (10, "ret10")]:
        cutoff = str((last_date - pd.DateOffset(years=years)).date())
        sub = prices[prices["date"] >= cutoff]
        if not sub.empty:
            start = float(sub["close"].iloc[0])
            if start > 0:
                out[key] = f"{(latest - start) / start * 100:+.1f}%"

    yr1 = prices[prices["date"] >= str((last_date - pd.DateOffset(years=1)).date())]
    if not yr1.empty and "high" in yr1.columns and "low" in yr1.columns:
        out["w52h"] = f"${yr1['high'].max():,.2f}"
        out["w52l"] = f"${yr1['low'].min():,.2f}"
    return out


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_top10_excel(companies_df: pd.DataFrame) -> bytes:
    """
    Generates a styled 4-sheet Excel report for the Top 10 ML-ranked stocks.
    Sheets: Cover | Top 10 Rankings | Fundamental Trends | Price Performance
    Returns raw bytes for st.download_button.
    """
    top10 = db.get_ml_recommendations().head(10)
    wb    = openpyxl.Workbook()

    # ── Sheet 0 : Cover ───────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Cover"
    ws0["A1"] = "FinMine Analytics Engine"
    ws0["A1"].font = Font(size=22, bold=True, color="1F3864")
    ws0["A2"] = "Top 10 Fundamental ML Investment Report"
    ws0["A2"].font = Font(size=14, bold=True, color="2F5496")
    ws0["A3"] = ""
    ws0["A4"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws0["A4"].font = Font(size=11)
    ws0["A5"] = f"Stocks in database: {len(companies_df):,}"
    ws0["A6"] = f"ML recommendations: {len(db.get_ml_recommendations()):,}"
    ws0["A7"] = ""
    ws0["A8"] = "Colour legend:"
    ws0["A8"].font = Font(bold=True)
    for row, (text, fill) in enumerate([
        ("STRONG BUY / ALPHACLASS  (score ≥ 0.72)", _SB_FILL),
        ("BUY                      (score ≥ 0.55)", _BUY_FILL),
        ("HOLD / UNDERPERFORM      (score  < 0.55)", _HOLD_FILL),
    ], 9):
        c = ws0.cell(row=row, column=1, value=text)
        c.fill = fill
        c.font = Font(size=11)
    ws0["A13"] = "For educational and research purposes only. Not investment advice."
    ws0["A13"].font = Font(italic=True, color="808080")
    ws0.column_dimensions["A"].width = 55

    # ── Sheet 1 : Top 10 Rankings ─────────────────────────────────────────
    ws1 = wb.create_sheet("Top 10 Rankings")
    rank_headers = [
        "Rank", "Ticker", "Company Name", "Sector",
        "ML Score", "ML Label",
        "Rev CAGR", "NI CAGR", "FCF CAGR",
        "Avg Op Margin", "Avg FCF Margin",
        "Avg D/E", "Cash Stability",
        "Latest Price", "1yr Return", "3yr Return", "10yr Return",
    ]
    for ci, h in enumerate(rank_headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws1.row_dimensions[1].height = 30

    for ri, (_, rec) in enumerate(top10.iterrows(), 2):
        ticker = rec["ticker"]
        comp   = companies_df[companies_df["ticker"] == ticker]
        ml_full   = db.get_ml_recommendation(ticker)
        features  = (ml_full or {}).get("features", {})
        pstats    = _price_stats(ticker)

        row_vals = [
            ri - 1,
            ticker,
            comp["name"].values[0] if not comp.empty else "",
            comp["sector"].values[0] if not comp.empty else "",
            round(float(rec["score"]), 4),
            rec["label"],
            _safe_pct(features.get("revenue_cagr")),
            _safe_pct(features.get("net_income_cagr")),
            _safe_pct(features.get("fcf_cagr")),
            _safe_pct(features.get("avg_op_margin")),
            _safe_pct(features.get("avg_fcf_margin")),
            _safe_ratio(features.get("avg_de_ratio")),
            _safe_ratio(features.get("avg_cash_stability")),
            pstats["latest"],
            pstats["ret1"],
            pstats["ret3"],
            pstats["ret10"],
        ]
        fill = _label_fill(rec["label"])
        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.fill = fill
            if ci == 1:
                c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws1, [6,10,30,22,10,26,10,10,10,13,13,10,14,14,12,12,12])

    # ── Sheet 2 : Fundamental Trends ──────────────────────────────────────
    ws2  = wb.create_sheet("Fundamental Trends")
    wr   = 1
    fund_headers = ["Period", "Revenue ($B)", "Net Income ($B)", "FCF ($B)",
                    "Op Margin %", "FCF Margin %", "D/E Ratio", "Cash ($B)"]

    for _, rec in top10.iterrows():
        ticker = rec["ticker"]
        comp   = companies_df[companies_df["ticker"] == ticker]
        cname  = comp["name"].values[0] if not comp.empty else ticker

        # Ticker section header
        title_cell = ws2.cell(
            row=wr, column=1,
            value=f"{ticker} — {cname}  |  ML Score: {rec['score']:.4f}  |  {rec['label']}"
        )
        title_cell.fill = _SUB_FILL
        title_cell.font = _SUB_FONT
        ws2.merge_cells(start_row=wr, start_column=1,
                        end_row=wr, end_column=len(fund_headers))
        wr += 1

        for ci, h in enumerate(fund_headers, 1):
            c = ws2.cell(row=wr, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.font = Font(bold=True)
        wr += 1

        ann_df = db.get_fundamentals(ticker, period_type="annual")
        if not ann_df.empty:
            ann_df["data"] = ann_df["data_json"].apply(
                lambda x: json.loads(x) if isinstance(x, str) else x
            )

            def _gseries(stmt: str, key: str) -> pd.Series:
                sub = ann_df[ann_df["statement_type"] == stmt]
                vals: dict = {}
                for _, row in sub.iterrows():
                    for k, v in row["data"].items():
                        if (k.lower().replace(" ", "").replace("_", "") ==
                                key.lower().replace(" ", "").replace("_", "")):
                            try:
                                vals[row["period_end"][:7]] = float(v)
                            except (TypeError, ValueError):
                                pass
                            break
                return pd.Series(vals).sort_index()

            rev  = _gseries("income",   "TotalRevenue")
            ni   = _gseries("income",   "NetIncome")
            oi   = _gseries("income",   "OperatingIncome")
            fcf  = _gseries("cashflow", "FreeCashFlow")
            debt = _gseries("balance",  "TotalDebt")
            cash = _gseries("balance",  "CashAndCashEquivalents")
            eq   = _gseries("balance",  "StockholdersEquity")

            all_periods = sorted(set(rev.index) | set(ni.index) | set(fcf.index))
            for period in all_periods[-10:]:
                r = rev.get(period);  n = ni.get(period)
                f = fcf.get(period);  o = oi.get(period)
                d = debt.get(period); c_val = cash.get(period)
                e = eq.get(period)
                op_m  = (o / r * 100) if r and r != 0 and o is not None else None
                fcf_m = (f / r * 100) if r and r != 0 and f is not None else None
                de    = (d / e)        if e and e != 0 and d is not None else None

                row_vals = [
                    period,
                    round(r / 1e9, 3)   if r       is not None else "",
                    round(n / 1e9, 3)   if n       is not None else "",
                    round(f / 1e9, 3)   if f       is not None else "",
                    f"{op_m:.1f}%"      if op_m    is not None else "",
                    f"{fcf_m:.1f}%"     if fcf_m   is not None else "",
                    f"{de:.2f}x"        if de      is not None else "",
                    round(c_val / 1e9, 3) if c_val is not None else "",
                ]
                for ci, val in enumerate(row_vals, 1):
                    ws2.cell(row=wr, column=ci, value=val)
                wr += 1

        wr += 2   # blank rows between companies

    _set_col_widths(ws2, [12, 16, 16, 12, 12, 12, 12, 12])

    # ── Sheet 3 : Price Performance ───────────────────────────────────────
    ws3 = wb.create_sheet("Price Performance")
    price_headers = [
        "Rank", "Ticker", "Company", "Sector",
        "Latest Price", "52W High", "52W Low",
        "1yr Return", "3yr Return", "10yr Return",
        "ML Score", "ML Label",
    ]
    for ci, h in enumerate(price_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for ri, (_, rec) in enumerate(top10.iterrows(), 2):
        ticker = rec["ticker"]
        comp   = companies_df[companies_df["ticker"] == ticker]
        ps     = _price_stats(ticker)
        fill   = _label_fill(rec["label"])

        row_vals = [
            ri - 1,
            ticker,
            comp["name"].values[0]   if not comp.empty else "",
            comp["sector"].values[0] if not comp.empty else "",
            ps["latest"], ps["w52h"], ps["w52l"],
            ps["ret1"], ps["ret3"], ps["ret10"],
            round(float(rec["score"]), 4),
            rec["label"],
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill = fill

    _set_col_widths(ws3, [6, 10, 30, 22, 14, 14, 14, 14, 14, 14, 10, 26])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_db_csv_zip(include_prices: bool = False) -> bytes:
    """
    Creates a ZIP archive containing every database table as a CSV.
    historical_prices is opt-in because it can be very large.
    Returns raw bytes for st.download_button.
    """
    buf = io.BytesIO()
    ts  = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        with db.get_conn() as conn:
            # companies
            df = pd.read_sql("SELECT * FROM companies ORDER BY ticker", conn)
            zf.writestr("companies.csv", df.to_csv(index=False))

            # ml_recommendations — expand features_json into a flat column
            df = pd.read_sql(
                "SELECT ticker, score, label, compiled_at, features_json "
                "FROM ml_recommendations ORDER BY score DESC", conn
            )
            zf.writestr("ml_recommendations.csv", df.to_csv(index=False))

            # financial_fundamentals — keep data_json as-is (schema varies per row)
            df = pd.read_sql(
                "SELECT ticker, period_type, period_end, statement_type, data_json, updated_at "
                "FROM financial_fundamentals ORDER BY ticker, period_end", conn
            )
            zf.writestr("financial_fundamentals.csv", df.to_csv(index=False))

            # ai_summaries
            df = pd.read_sql(
                "SELECT ticker, model, created_at, markdown "
                "FROM ai_summaries ORDER BY ticker", conn
            )
            zf.writestr("ai_summaries.csv", df.to_csv(index=False))

            # mining_checkpoint
            df = pd.read_sql(
                "SELECT ticker, status, error_msg, updated_at "
                "FROM mining_checkpoint ORDER BY ticker", conn
            )
            zf.writestr("mining_checkpoint.csv", df.to_csv(index=False))

            if include_prices:
                df = pd.read_sql(
                    "SELECT ticker, date, open, high, low, close, volume, dividends, splits "
                    "FROM historical_prices ORDER BY ticker, date", conn
                )
                zf.writestr("historical_prices.csv", df.to_csv(index=False))

        readme = (
            "FinMine Analytics Engine — Database Export\n"
            "==========================================\n"
            f"Generated: {ts}\n\n"
            "Files included:\n"
            "  companies.csv             — All tracked companies\n"
            "  ml_recommendations.csv    — ML confidence scores and labels\n"
            "  financial_fundamentals.csv — Annual/quarterly statement data (JSON field)\n"
            "  ai_summaries.csv          — Cached Claude AI analysis\n"
            "  mining_checkpoint.csv     — Pipeline status per ticker\n"
        )
        if include_prices:
            readme += "  historical_prices.csv     — Daily OHLCV data\n"
        readme += "\nFor educational and research purposes only.\n"
        zf.writestr("README.txt", readme)

    buf.seek(0)
    return buf.getvalue()


def single_table_csv(table: str, order_by: str = "1") -> bytes:
    """Exports a single DB table to CSV bytes."""
    with db.get_conn() as conn:
        df = pd.read_sql(f"SELECT * FROM {table} ORDER BY {order_by}", conn)
    return df.to_csv(index=False).encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — API KEY + DB STATS
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/combo-chart.png", width=56)
    st.title("FinMine Analytics")
    st.caption("Production-grade equity research engine")
    st.divider()

    api_key_input = st.text_input(
        "Anthropic API Key",
        type="password",
        placeholder="sk-ant-...",
        help="Required only for the AI Deep-Dive feature.",
    )

    st.divider()
    st.subheader("Database Status")
    stats = db.db_stats()
    c1, c2 = st.columns(2)
    c1.metric("Companies",    f"{stats['companies']:,}")
    c2.metric("Price Rows",   f"{stats['historical_prices']:,}")
    c1.metric("Fundamentals", f"{stats['financial_fundamentals']:,}")
    c2.metric("ML Recs",      f"{stats['ml_recommendations']:,}")
    c1.metric("AI Summaries", f"{stats['ai_summaries']:,}")
    c2.metric("DB Size",      f"{stats['db_size_mb']} MB")

    st.divider()
    st.caption("All data is stored locally in `finmine.db`. "
               "No data leaves your machine except for AI summary calls.")


# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────

tab_mine, tab_stock, tab_ml, tab_export, tab_help = st.tabs([
    "⛏️ Data Mine",
    "📊 Stock Analysis",
    "🤖 ML Rankings",
    "📥 Export & Reports",
    "❓ Help",
])


# ═════════════════════════════════════════════════════════════════════════════
# TAB 1 — DATA MINE
# ═════════════════════════════════════════════════════════════════════════════

with tab_mine:
    st.header("⛏️ Data Mining Engine")
    st.caption(
        "Download 10 years of OHLCV + financial statements for the full US equity universe "
        "or just the S&P 500. Mining resumes from checkpoint if interrupted."
    )

    col_a, col_b = st.columns([1, 2])
    with col_a:
        universe_mode = st.radio(
            "Universe",
            ["S&P 500 only (~500 tickers, ~10 min)",
             "Full US equity universe (~7,000 tickers, several hours)"],
            index=0,
        )
        run_ml_after = st.checkbox("Run ML pipeline automatically after mine", value=True)
        force_refresh_mine = st.checkbox("Force re-download even if data exists", value=False)

    with col_b:
        st.info(
            "**Checkpoint system:** If mining is interrupted, click 'Start Mine' again — "
            "it will skip tickers already marked `done` and resume where it left off. "
            "No duplicate API calls are made."
        )

    start_mine_btn = st.button("Start Data Mine", type="primary", key="start_mine")

    if start_mine_btn:
        use_sp500 = "S&P 500" in universe_mode

        log_container = st.empty()
        prog_price    = st.progress(0, text="Prices: waiting...")
        prog_fund     = st.progress(0, text="Fundamentals: waiting...")
        status_box    = st.empty()
        log_lines: list[str] = []

        def _log_cb(msg: str):
            log_lines.append(f"[{datetime.utcnow().strftime('%H:%M:%S')}] {msg}")
            log_container.code("\n".join(log_lines[-20:]))

        price_total = {"n": 1}
        price_done  = {"n": 0}

        def _price_cb(ticker: str):
            price_done["n"] += 1
            pct = min(price_done["n"] / max(price_total["n"], 1), 1.0)
            prog_price.progress(pct, text=f"Prices: {price_done['n']}/{price_total['n']} — {ticker}")

        def _fund_cb(ticker: str, done: int, total: int):
            pct = min(done / max(total, 1), 1.0)
            prog_fund.progress(pct, text=f"Fundamentals: {done}/{total} — {ticker}")

        with st.spinner("Mining in progress..."):
            try:
                universe = scraper.get_ticker_universe(use_sp500_only=use_sp500)
                price_total["n"] = len(universe)

                if force_refresh_mine:
                    for u in universe:
                        db.checkpoint_set(u["ticker"], "pending")

                mine_stats = scraper.run_full_mine(
                    use_sp500_only=use_sp500,
                    price_progress_cb=_price_cb,
                    fund_progress_cb=_fund_cb,
                    log_cb=_log_cb,
                )
                status_box.success(
                    f"Mine complete — {mine_stats['universe_size']} tickers | "
                    f"Prices OK: {mine_stats['prices_ok']} | "
                    f"Fundamentals OK: {mine_stats['funds_ok']}"
                )
            except Exception as exc:
                status_box.error(f"Mine failed: {exc}")
                st.stop()

        if run_ml_after:
            ml_log_lines: list[str] = []
            ml_log_box   = st.empty()
            ml_prog      = st.progress(0, text="ML: building features...")

            def _feat_cb(ticker: str, done: int, total: int):
                pct = min(done / max(total, 1), 1.0)
                ml_prog.progress(pct, text=f"ML features: {done}/{total} — {ticker}")

            def _ml_log(msg: str):
                ml_log_lines.append(f"[{datetime.utcnow().strftime('%H:%M:%S')}] {msg}")
                ml_log_box.code("\n".join(ml_log_lines[-10:]))

            with st.spinner("Running ML pipeline..."):
                ml_result = ml_engine.run_ml_pipeline(
                    feat_progress_cb=_feat_cb,
                    log_cb=_ml_log,
                )
            if ml_result.get("status") == "ok":
                st.success(
                    f"ML complete — {ml_result['tickers_scored']} stocks scored | "
                    f"STRONG BUY: {ml_result['strong_buy']} | "
                    f"BUY: {ml_result['buy']} | "
                    f"HOLD: {ml_result['hold']}"
                )
            else:
                st.warning(f"ML pipeline: {ml_result}")

        st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# TAB 2 — STOCK ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════

with tab_stock:
    st.header("📊 Fundamental Stock Analysis")

    companies_df = db.get_all_companies()
    all_tickers  = sorted(companies_df["ticker"].tolist()) if not companies_df.empty else []

    col_s1, col_s2 = st.columns([1, 3])
    with col_s1:
        if all_tickers:
            default_idx   = all_tickers.index("AAPL") if "AAPL" in all_tickers else 0
            chosen_ticker = st.selectbox("Select Ticker", all_tickers, index=default_idx)
        else:
            chosen_ticker = st.text_input("Ticker (type to search)", "AAPL").upper()

        refresh_btn = st.button("Refresh from API", help="Force re-download even if cached")

    with col_s2:
        if chosen_ticker:
            comp_row = (companies_df[companies_df["ticker"] == chosen_ticker]
                        if not companies_df.empty else pd.DataFrame())
            if not comp_row.empty:
                r = comp_row.iloc[0]
                st.markdown(f"### {r.get('name', chosen_ticker)}")
                st.caption(f"{r.get('exchange','')} | {r.get('sector','')} | {r.get('industry','')}")
            else:
                st.markdown(f"### {chosen_ticker}")

    if not chosen_ticker:
        st.info("Select or type a ticker to begin analysis.")
        st.stop()

    if refresh_btn or not db.company_exists(chosen_ticker):
        with st.spinner(f"Fetching data for {chosen_ticker}..."):
            status = scraper.refresh_ticker(chosen_ticker, force=refresh_btn)
        col_s1.caption(f"Prices: {status['prices']} | Fundamentals: {status['fundamentals']}")

    prices_df = db.get_prices(chosen_ticker)
    ann_df    = db.get_fundamentals(chosen_ticker, period_type="annual")
    ml_rec    = db.get_ml_recommendation(chosen_ticker)

    if prices_df.empty:
        st.warning(f"No price data found for {chosen_ticker}. Click 'Refresh from API'.")
        st.stop()

    prices_df = prices_df.sort_values("date")
    prices_df["date"] = pd.to_datetime(prices_df["date"])

    if ml_rec:
        label = ml_rec["label"]
        score = ml_rec["score"]
        color = {"STRONG BUY / ALPHACLASS": "🟢", "BUY": "🔵",
                 "HOLD / UNDERPERFORM": "🟡"}.get(label, "⚪")
        st.info(f"**ML Engine:** {color} {label} — Confidence Score: **{score:.3f}**  "
                f"_(compiled {ml_rec.get('compiled_at','')[:10]})_")

    m1, m2, m3, m4, m5 = st.columns(5)
    latest_close   = float(prices_df["close"].iloc[-1])
    earliest_close = float(prices_df["close"].iloc[0])
    m1.metric("Latest Close",  f"${latest_close:,.2f}")
    m2.metric("10-yr Start",   f"${earliest_close:,.2f}")
    pct_10yr = (latest_close - earliest_close) / earliest_close * 100
    m3.metric("10-yr Return",  f"{pct_10yr:+.1f}%")
    m4.metric("Price Rows",    f"{len(prices_df):,}")
    m5.metric("Data From",     str(prices_df["date"].min())[:10])

    st.divider()

    if not ann_df.empty:
        ann_df["data"] = ann_df["data_json"].apply(
            lambda x: json.loads(x) if isinstance(x, str) else x
        )

        def get_annual_series(stmt_type: str, key: str) -> pd.Series:
            sub  = ann_df[ann_df["statement_type"] == stmt_type]
            vals = {}
            for _, row in sub.iterrows():
                for k, v in row["data"].items():
                    if (k.lower().replace(" ", "").replace("_", "") ==
                            key.lower().replace(" ", "").replace("_", "")):
                        try:
                            vals[row["period_end"][:7]] = float(v)
                        except (TypeError, ValueError):
                            pass
                        break
            return pd.Series(vals).sort_index()

        revenue    = get_annual_series("income",   "TotalRevenue")
        net_income = get_annual_series("income",   "NetIncome")
        op_income  = get_annual_series("income",   "OperatingIncome")
        total_debt = get_annual_series("balance",  "TotalDebt")
        total_cash = get_annual_series("balance",  "CashAndCashEquivalents")
        equity_s   = get_annual_series("balance",  "StockholdersEquity")
        fcf        = get_annual_series("cashflow", "FreeCashFlow")
        capex      = get_annual_series("cashflow", "CapitalExpenditure")

        op_margin = (op_income / revenue.reindex(op_income.index)).dropna() * 100
        de_ratio  = (total_debt / equity_s.reindex(total_debt.index)
                     .replace(0, np.nan)).dropna()

        with st.expander("Chart 1 — Revenue, Net Income & Operating Margin", expanded=True):
            fig1 = sp.make_subplots(specs=[[{"secondary_y": True}]])
            if not revenue.empty:
                fig1.add_trace(go.Bar(x=revenue.index, y=revenue.values / 1e9,
                    name="Revenue ($B)", marker_color="steelblue", opacity=0.75),
                    secondary_y=False)
            if not net_income.empty:
                fig1.add_trace(go.Scatter(x=net_income.index, y=net_income.values / 1e9,
                    name="Net Income ($B)", line=dict(color="limegreen", width=2)),
                    secondary_y=False)
            if not op_margin.empty:
                fig1.add_trace(go.Scatter(x=op_margin.index, y=op_margin.values,
                    name="Operating Margin (%)", line=dict(color="orange", width=2, dash="dot")),
                    secondary_y=True)
            fig1.update_layout(
                title=f"{chosen_ticker} — Revenue, Net Income & Operating Margin (10-year)",
                hovermode="x unified", height=420, legend=dict(orientation="h", y=-0.15))
            fig1.update_yaxes(title_text="$ Billions", secondary_y=False)
            fig1.update_yaxes(title_text="Operating Margin (%)", secondary_y=True)
            st.plotly_chart(fig1, use_container_width=True)

        with st.expander("Chart 2 — Debt/Equity Ratio vs. Cash Reserves", expanded=True):
            fig2 = sp.make_subplots(specs=[[{"secondary_y": True}]])
            if not de_ratio.empty:
                fig2.add_trace(go.Scatter(x=de_ratio.index, y=de_ratio.values,
                    name="Debt-to-Equity", line=dict(color="tomato", width=2),
                    fill="tozeroy", fillcolor="rgba(255,99,71,0.1)"), secondary_y=False)
            if not total_cash.empty:
                fig2.add_trace(go.Bar(x=total_cash.index, y=total_cash.values / 1e9,
                    name="Cash & Equivalents ($B)", marker_color="mediumseagreen", opacity=0.7),
                    secondary_y=True)
            fig2.update_layout(
                title=f"{chosen_ticker} — Debt/Equity Trajectory vs. Cash Reserves",
                hovermode="x unified", height=400, legend=dict(orientation="h", y=-0.15))
            fig2.update_yaxes(title_text="D/E Ratio (x)", secondary_y=False)
            fig2.update_yaxes(title_text="Cash ($B)", secondary_y=True)
            st.plotly_chart(fig2, use_container_width=True)

        with st.expander("Chart 3 — Free Cash Flow vs. Capital Expenditures", expanded=True):
            fig3 = go.Figure()
            if not fcf.empty:
                fig3.add_trace(go.Bar(x=fcf.index, y=fcf.values / 1e9,
                    name="Free Cash Flow ($B)", marker_color="mediumslateblue", opacity=0.8))
            if not capex.empty:
                fig3.add_trace(go.Bar(x=capex.abs().index, y=-capex.abs().values / 1e9,
                    name="CapEx ($B, inverted)", marker_color="salmon", opacity=0.8))
            fig3.update_layout(
                title=f"{chosen_ticker} — Free Cash Flow vs. CapEx (10-year)",
                barmode="overlay", hovermode="x unified", height=400,
                yaxis_title="$ Billions", legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig3, use_container_width=True)
    else:
        st.warning("No annual fundamental data available. Run the Data Mine or click Refresh.")

    with st.expander("Chart 4 — Historical Price", expanded=True):
        fig4 = go.Figure()
        fig4.add_trace(go.Scatter(x=prices_df["date"], y=prices_df["close"],
            name="Adj. Close", line=dict(color="dodgerblue", width=1.5),
            fill="tozeroy", fillcolor="rgba(30,144,255,0.06)"))
        if not prices_df["volume"].isna().all():
            fig4.add_trace(go.Bar(x=prices_df["date"], y=prices_df["volume"],
                name="Volume", marker_color="rgba(100,100,200,0.25)", yaxis="y2"))
        fig4.update_layout(
            title=f"{chosen_ticker} — 10-year Adjusted Close Price",
            hovermode="x unified", height=450,
            xaxis=dict(rangeslider=dict(visible=True)),
            yaxis=dict(title="Price ($)"),
            yaxis2=dict(title="Volume", overlaying="y", side="right",
                        showgrid=False, tickformat=".2s"),
            legend=dict(orientation="h", y=-0.15))
        st.plotly_chart(fig4, use_container_width=True)

    st.divider()
    st.subheader("🤖 AI Fundamental Deep-Dive")

    cached_summary = db.get_ai_summary(chosen_ticker)
    if cached_summary:
        st.caption(f"Cached summary from {cached_summary.get('created_at','')[:10]} "
                   f"(model: {cached_summary.get('model','')})")
        st.markdown(cached_summary["markdown"])
        if st.button("Regenerate AI Summary", key="regen_ai"):
            if not api_key_input:
                st.error("Enter your Anthropic API key in the sidebar.")
            else:
                with st.spinner("Calling Claude API..."):
                    try:
                        text = ai_engine.generate_ai_summary(
                            chosen_ticker, api_key_input, force_refresh=True)
                        st.markdown(text)
                        st.success("Summary regenerated and cached.")
                    except Exception as exc:
                        st.error(f"API error: {exc}")
    else:
        st.info("No cached summary. Click the button below to generate one.")
        if st.button("🤖 Generate AI Fundamental Deep-Dive", type="primary", key="gen_ai"):
            if not api_key_input:
                st.error("Enter your Anthropic API key in the sidebar.")
            else:
                with st.spinner(f"Asking Claude to analyse {chosen_ticker}..."):
                    try:
                        text = ai_engine.generate_ai_summary(chosen_ticker, api_key_input)
                        st.markdown(text)
                        st.success("Summary saved to local database.")
                    except Exception as exc:
                        st.error(f"API error: {exc}")

    with st.expander("Raw fundamental data tables"):
        if not ann_df.empty:
            for stmt in ("income", "balance", "cashflow"):
                sub = ann_df[ann_df["statement_type"] == stmt][["period_end", "data"]]
                if sub.empty:
                    continue
                st.markdown(f"**{stmt.title()} Statement (Annual)**")
                rows = []
                for _, row in sub.iterrows():
                    flat = {"period_end": row["period_end"]}
                    flat.update(row["data"] if isinstance(row["data"], dict)
                                else json.loads(row.get("data_json", "{}")))
                    rows.append(flat)
                st.dataframe(pd.DataFrame(rows).set_index("period_end"),
                             use_container_width=True)
        else:
            st.info("No fundamentals in database for this ticker.")


# ═════════════════════════════════════════════════════════════════════════════
# TAB 3 — ML RANKINGS
# ═════════════════════════════════════════════════════════════════════════════

with tab_ml:
    st.header("🤖 ML Predictive Rankings")
    st.caption(
        "Random Forest + Gradient Boosting ensemble. "
        "Scores reflect probability of fundamental outperformance vs. market median. "
        "Updated each time the Data Mine completes."
    )

    recs_df = db.get_ml_recommendations()

    if recs_df.empty:
        st.info("No ML recommendations yet. Run the Data Mine (with ML pipeline enabled) first.")
    else:
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            min_score = st.slider("Minimum Confidence Score", 0.0, 1.0, 0.0, 0.05)
        with col_f2:
            label_filter = st.multiselect(
                "Labels",
                ["STRONG BUY / ALPHACLASS", "BUY", "HOLD / UNDERPERFORM"],
                default=["STRONG BUY / ALPHACLASS", "BUY", "HOLD / UNDERPERFORM"],
            )
        with col_f3:
            top_n = st.selectbox("Top N to display", [25, 50, 100, 250, "All"], index=1)

        filtered = recs_df[recs_df["score"] >= min_score]
        if label_filter:
            filtered = filtered[filtered["label"].isin(label_filter)]
        if top_n != "All":
            filtered = filtered.head(int(top_n))

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Stocks Scored",     f"{len(recs_df):,}")
        mc2.metric("STRONG BUY",        f"{(recs_df['label']=='STRONG BUY / ALPHACLASS').sum():,}")
        mc3.metric("BUY",               f"{(recs_df['label']=='BUY').sum():,}")
        mc4.metric("HOLD/UNDERPERFORM", f"{(recs_df['label']=='HOLD / UNDERPERFORM').sum():,}")

        with st.expander("Score Distribution", expanded=False):
            fig_dist = go.Figure(go.Histogram(
                x=recs_df["score"], nbinsx=40,
                marker_color="steelblue", opacity=0.8, name="Confidence Score"))
            fig_dist.add_vline(x=0.72, line_dash="dash", line_color="limegreen",
                               annotation_text="STRONG BUY (0.72)")
            fig_dist.add_vline(x=0.55, line_dash="dash", line_color="dodgerblue",
                               annotation_text="BUY (0.55)")
            fig_dist.update_layout(title="ML Confidence Score Distribution",
                                   xaxis_title="Score", yaxis_title="Count", height=320)
            st.plotly_chart(fig_dist, use_container_width=True)

        disp = filtered[["ticker", "score", "label", "compiled_at"]].copy()
        disp["score"] = disp["score"].round(4)
        disp["rank"]  = range(1, len(disp) + 1)
        disp = disp[["rank", "ticker", "score", "label", "compiled_at"]]

        def _color_label(v: str) -> str:
            if "STRONG" in v:
                return "color: limegreen; font-weight: bold"
            if v == "BUY":
                return "color: dodgerblue"
            return "color: orange"

        styled = disp.style.map(_color_label, subset=["label"])
        st.dataframe(styled, use_container_width=True, hide_index=True, height=500)

        with st.expander("Top 20 by Confidence Score", expanded=True):
            top20  = filtered.head(20)
            colors = [
                "limegreen" if "STRONG" in lbl else
                "dodgerblue" if lbl == "BUY" else "orange"
                for lbl in top20["label"]
            ]
            fig_bar = go.Figure(go.Bar(
                x=top20["ticker"], y=top20["score"],
                marker_color=colors,
                text=[f"{s:.3f}" for s in top20["score"]],
                textposition="outside",
            ))
            fig_bar.update_layout(
                title="Top 20 ML-Ranked Equities",
                xaxis_title="Ticker", yaxis_title="Confidence Score",
                yaxis=dict(range=[0, 1.05]), height=420)
            st.plotly_chart(fig_bar, use_container_width=True)

        st.subheader("Feature Breakdown")
        ml_ticker_sel = st.selectbox("Inspect ticker features",
                                     filtered["ticker"].tolist(), key="ml_feat_sel")
        if ml_ticker_sel:
            rec = db.get_ml_recommendation(ml_ticker_sel)
            if rec and rec.get("features"):
                feat_df = pd.DataFrame(
                    [(k, v) for k, v in rec["features"].items() if v is not None],
                    columns=["Feature", "Value"],
                )
                feat_df["Value"] = feat_df["Value"].apply(
                    lambda v: f"{float(v):.4f}" if v is not None else "N/A")
                st.dataframe(feat_df, use_container_width=True, hide_index=True)
            else:
                st.info("No feature data stored for this ticker.")

        # ── Quick Top 10 download shortcut ────────────────────────────────
        st.divider()
        st.markdown("##### Quick Export")
        companies_df_ml = db.get_all_companies()
        if st.button("Generate Top 10 Excel Report", key="top10_quick"):
            with st.spinner("Building report..."):
                try:
                    xlsx_bytes = build_top10_excel(companies_df_ml)
                    fname = f"FinMine_Top10_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
                    st.download_button(
                        label="📥 Download Top 10 Report (.xlsx)",
                        data=xlsx_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="top10_dl_quick",
                    )
                except Exception as exc:
                    st.error(f"Report generation failed: {exc}")

        st.divider()
        if st.button("Re-run ML Pipeline (uses existing DB data)", key="run_ml_standalone"):
            ml_log   = st.empty()
            ml_prog2 = st.progress(0)
            ml_lines: list[str] = []

            def _fc2(ticker: str, done: int, total: int):
                ml_prog2.progress(min(done / max(total, 1), 1.0),
                                  text=f"Features: {done}/{total}")

            def _ml2(msg: str):
                ml_lines.append(msg)
                ml_log.code("\n".join(ml_lines[-8:]))

            with st.spinner("Running ML pipeline..."):
                res = ml_engine.run_ml_pipeline(feat_progress_cb=_fc2, log_cb=_ml2)
            if res.get("status") == "ok":
                st.success(f"ML complete: {res['tickers_scored']} scored")
                st.rerun()
            else:
                st.warning(f"ML result: {res}")


# ═════════════════════════════════════════════════════════════════════════════
# TAB 4 — EXPORT & REPORTS
# ═════════════════════════════════════════════════════════════════════════════

with tab_export:
    st.header("📥 Export & Reports")
    st.caption("Download investment reports and database exports to your local machine.")

    companies_df_exp = db.get_all_companies()
    recs_exp         = db.get_ml_recommendations()

    # ── Section 1: Top 10 Investment Report ──────────────────────────────
    st.subheader("🏆 Top 10 Investment Report (.xlsx)")
    st.markdown(
        "Generates a **styled 4-sheet Excel workbook** covering the top 10 "
        "ML-ranked stocks. Includes fundamental trends, price performance, "
        "and ML feature breakdown — colour-coded by label."
    )

    col_r1, col_r2 = st.columns([2, 1])
    with col_r1:
        st.markdown("""
**Sheets included:**
| Sheet | Contents |
|-------|---------|
| Cover | Report metadata, colour legend |
| Top 10 Rankings | Rank, ticker, company, ML score, label, all key metrics, price returns |
| Fundamental Trends | Annual revenue / NI / FCF / margins / D/E for each stock (10-year) |
| Price Performance | Latest price, 52W high/low, 1yr / 3yr / 10yr returns |
""")
    with col_r2:
        if recs_exp.empty:
            st.warning("No ML recommendations in database yet.\n\nRun the Data Mine first.")
        else:
            top10_preview = recs_exp.head(10)[["ticker", "score", "label"]].copy()
            top10_preview["score"] = top10_preview["score"].round(4)
            top10_preview.index = range(1, len(top10_preview) + 1)
            st.dataframe(top10_preview, use_container_width=True)

    if not recs_exp.empty:
        if st.button("Build Top 10 Excel Report", type="primary", key="build_top10"):
            with st.spinner("Assembling report — fetching fundamentals and prices..."):
                try:
                    xlsx_bytes = build_top10_excel(companies_df_exp)
                    fname = f"FinMine_Top10_Report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
                    st.download_button(
                        label="📥 Download Top 10 Report (.xlsx)",
                        data=xlsx_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="top10_download",
                    )
                    st.success(f"Report ready: **{fname}**  ({len(xlsx_bytes)/1024:.0f} KB)")
                except Exception as exc:
                    st.error(f"Report failed: {exc}")
                    log.exception("Top 10 report error")

    st.divider()

    # ── Section 2: Database CSV Exports ───────────────────────────────────
    st.subheader("🗃️ Full Database Export (.zip)")
    st.markdown(
        "Downloads all database tables as CSV files bundled in a single ZIP archive. "
        "Use this to analyse the data in Excel, Python, or any BI tool."
    )

    col_z1, col_z2 = st.columns([2, 1])
    with col_z1:
        include_prices_chk = st.checkbox(
            "Include historical_prices.csv",
            value=False,
            help=f"Price table has {stats['historical_prices']:,} rows and can be very large.",
        )
        st.caption(
            "Files always included: `companies`, `ml_recommendations`, "
            "`financial_fundamentals`, `ai_summaries`, `mining_checkpoint`, `README.txt`"
        )
    with col_z2:
        st.metric("Rows in DB", f"{sum(v for k,v in stats.items() if k != 'db_size_mb'):,}")
        st.metric("DB file size", f"{stats['db_size_mb']} MB")

    if st.button("Build Full Database ZIP", type="primary", key="build_zip"):
        with st.spinner("Compressing tables..."):
            try:
                zip_bytes = build_db_csv_zip(include_prices=include_prices_chk)
                fname_zip = f"FinMine_DB_Export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.zip"
                st.download_button(
                    label="📥 Download Database ZIP",
                    data=zip_bytes,
                    file_name=fname_zip,
                    mime="application/zip",
                    key="zip_download",
                )
                st.success(f"ZIP ready: **{fname_zip}**  ({len(zip_bytes)/1024:.0f} KB)")
            except Exception as exc:
                st.error(f"ZIP export failed: {exc}")
                log.exception("ZIP export error")

    st.divider()

    # ── Section 3: Individual Table Downloads ─────────────────────────────
    st.subheader("📋 Individual Table Downloads (.csv)")
    st.caption("Download a single table as a plain CSV file.")

    TABLE_CONFIG = {
        "companies":             ("Companies",              "ticker"),
        "ml_recommendations":    ("ML Recommendations",     "score DESC"),
        "financial_fundamentals":("Financial Fundamentals", "ticker, period_end"),
        "ai_summaries":          ("AI Summaries",           "ticker"),
        "mining_checkpoint":     ("Mining Checkpoint",      "ticker"),
    }

    cols_tables = st.columns(len(TABLE_CONFIG))
    for col, (table, (label, order)) in zip(cols_tables, TABLE_CONFIG.items()):
        with col:
            row_count = stats.get(table, 0)
            st.markdown(f"**{label}**")
            st.caption(f"{row_count:,} rows")
            if st.button(f"Download {label}", key=f"dl_{table}"):
                try:
                    csv_bytes = single_table_csv(table, order)
                    st.download_button(
                        label=f"📥 {table}.csv",
                        data=csv_bytes,
                        file_name=f"finmine_{table}_{datetime.utcnow().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        key=f"dl_btn_{table}",
                    )
                except Exception as exc:
                    st.error(f"Export failed: {exc}")

    st.divider()

    # ── Section 4: Top 10 as plain CSV ────────────────────────────────────
    st.subheader("📊 Top 10 Rankings — Plain CSV")
    st.caption(
        "A flat CSV with rank, ticker, company, sector, score, label, "
        "all ML features, and price returns for the top 10 stocks."
    )

    if not recs_exp.empty and not companies_df_exp.empty:
        top10_csv_rows = []
        for rank, (_, rec) in enumerate(recs_exp.head(10).iterrows(), 1):
            ticker   = rec["ticker"]
            comp_row = companies_df_exp[companies_df_exp["ticker"] == ticker]
            ml_full  = db.get_ml_recommendation(ticker)
            features = (ml_full or {}).get("features", {})
            ps       = _price_stats(ticker)

            row = {
                "rank":        rank,
                "ticker":      ticker,
                "company":     comp_row["name"].values[0]   if not comp_row.empty else "",
                "sector":      comp_row["sector"].values[0] if not comp_row.empty else "",
                "ml_score":    round(float(rec["score"]), 4),
                "ml_label":    rec["label"],
                "compiled_at": rec.get("compiled_at", ""),
                "latest_price":ps["latest"],
                "52w_high":    ps["w52h"],
                "52w_low":     ps["w52l"],
                "return_1yr":  ps["ret1"],
                "return_3yr":  ps["ret3"],
                "return_10yr": ps["ret10"],
            }
            for feat_key in ml_engine.FEATURE_COLS:
                row[feat_key] = features.get(feat_key, "")
            top10_csv_rows.append(row)

        if top10_csv_rows:
            top10_flat_df = pd.DataFrame(top10_csv_rows)
            csv_out = top10_flat_df.to_csv(index=False).encode("utf-8")
            fname_t10 = f"FinMine_Top10_{datetime.utcnow().strftime('%Y%m%d')}.csv"
            st.download_button(
                label="📥 Download Top 10 CSV",
                data=csv_out,
                file_name=fname_t10,
                mime="text/csv",
                key="top10_csv_dl",
            )
            st.dataframe(top10_flat_df[["rank","ticker","company","sector",
                                        "ml_score","ml_label","return_1yr",
                                        "return_3yr","return_10yr"]],
                         use_container_width=True, hide_index=True)
    else:
        st.info("Run the Data Mine and ML pipeline first to generate rankings.")


# ═════════════════════════════════════════════════════════════════════════════
# TAB 5 — HELP
# ═════════════════════════════════════════════════════════════════════════════

with tab_help:
    st.header("❓ FinMine Analytics Engine — User Guide")
    st.markdown("""
## Quick Start

1. **First run:** Go to the **⛏️ Data Mine** tab and click **Start Data Mine**.
   - Choose *S&P 500 only* for a ~10-minute initial run.
   - The full US equity universe (~7,000 tickers) takes several hours.
   - Mining is **checkpoint-based** — you can stop and restart at any time.

2. **Analyse a stock:** Go to **📊 Stock Analysis**, pick a ticker, and explore the 4 interactive charts.

3. **AI Deep-Dive:** Enter your Anthropic API key in the sidebar, then click **Generate AI Fundamental Deep-Dive**. The response is cached permanently.

4. **ML Rankings:** Go to **🤖 ML Rankings** to view the full sorted leaderboard.

5. **Export:** Go to **📥 Export & Reports** to download the Top 10 Excel report or full database CSVs.

---

## Architecture

| Module | Purpose |
|--------|---------|
| `database.py`    | SQLite WAL, schema, all read/write helpers |
| `scraper.py`     | Universe fetch, yfinance OHLCV + fundamentals, checkpoints |
| `ml_engine.py`   | Feature engineering, RF + GB ensemble, scoring |
| `ai_engine.py`   | Claude API integration, payload assembly, caching |
| `finmine_app.py` | Streamlit UI (this file) |

---

## Export & Reports Reference

| Export | Format | Contents |
|--------|--------|---------|
| Top 10 Excel Report | `.xlsx` (4 sheets) | Cover, Rankings, Fundamentals, Price Performance |
| Top 10 CSV | `.csv` | Flat file — all metrics + price returns |
| Full DB ZIP | `.zip` (5-6 CSVs) | All tables; prices optional |
| Individual tables | `.csv` each | One table at a time |

---

## ML Engine Details

- **Features:** Revenue CAGR, Net Income CAGR, FCF CAGR, Avg Operating Margin,
  FCF Margin, Margin velocity, D/E ratio & trend, Cash stability, scale metrics
- **Model:** Ensemble of Random Forest (300 trees) + Gradient Boosting (200 trees)
- **Target:** Binary — does the stock 3-year return beat universe median?
- **Labels:** STRONG BUY ≥ 0.72 | BUY ≥ 0.55 | HOLD < 0.55

---

## Staleness Policy

| Data type | Refreshed when |
|-----------|---------------|
| Prices | Older than 24 hours |
| Fundamentals | Older than 90 days |
| AI Summary | Manual regenerate only |
| ML Scores | After each Data Mine run |

---

## Troubleshooting

- **No data for ticker:** Click *Refresh from API* in Stock Analysis.
- **AI button grayed out:** Enter your `sk-ant-...` key in the sidebar.
- **Mine stalls:** Checkpoint-based — click *Start Data Mine* again to resume.
- **DB corruption:** Delete `finmine.db` and re-mine. All data is re-fetchable.
- **Top 10 report empty:** Run the ML pipeline first (Data Mine tab or ML Rankings tab).
""")


# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────

st.divider()
st.caption(
    "FinMine Analytics Engine — local financial data mining & AI research platform. "
    "For educational and research purposes only. Not investment advice."
)
